import json
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path(r"c:\Users\AmirRezaTM\Downloads\فرم برند نهایی ترین (1).xlsx")
OUT_PATH = Path(__file__).resolve().parents[1] / "app" / "lib" / "campaignSubTypes.ts"


def main() -> None:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["campaign types"]
    current_type = None
    mapping: dict[str, list[str]] = {}
    for row in range(2, ws.max_row + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value
        if col_a:
            current_type = str(col_a).strip()
        if col_b and current_type:
            mapping.setdefault(current_type, []).append(str(col_b).strip())

    content = f'''import {{ CAMPAIGN_TYPE_QUESTION }} from "./campaignKpis";

export const CAMPAIGN_SUBTYPE_QUESTION =
  "با توجه به این که کدام یک از دسته بندی ها را انتخاب میکنید، زیر دسته ی همان کمپین را مشخص کنید";

export {{ CAMPAIGN_TYPE_QUESTION }};

export const CAMPAIGN_SUBTYPE_MAP: Record<string, string[]> = {json.dumps(mapping, ensure_ascii=False, indent=2)};

export function getCampaignSubTypes(campaignType: string): string[] {{
  return CAMPAIGN_SUBTYPE_MAP[campaignType] ?? [];
}}
'''
    OUT_PATH.write_text(content, encoding="utf-8")
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
