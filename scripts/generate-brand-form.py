import json
import re
from pathlib import Path

from openpyxl import load_workbook

path = Path(r"c:\Users\AmirRezaTM\Downloads\فرم برند نهایی ترین.xlsx")
out_path = Path(__file__).resolve().parents[1] / "app" / "lib" / "brandForm.ts"
wb = load_workbook(path, data_only=True)
BRAND_SHEETS = [s for s in wb.sheetnames if s.startswith("صفحه")]
SKIP = {"سوال", "سوالات", "سوالات- توضیح در کامنت"}


def clean(value) -> str:
    return str(value).strip() if value else ""


def split_options(text: str):
    if not text or text.strip() in ("ندارد", "-", ""):
        return None
    if text.strip() in ("بله / خیر", "بله/خیر"):
        return ["بله", "خیر"]
    parts = re.split(r"\n\n+", text.strip())
    if len(parts) == 1:
        parts = [p.strip() for p in text.strip().split("\n") if p.strip()]
    else:
        parts = [p.strip() for p in parts if p.strip()]
    filtered = []
    for part in parts:
        if part.startswith("در صورتی که") or part.startswith("اگر در سوال"):
            continue
        if "نمایش داده" in part and len(part) > 80:
            continue
        filtered.append(part)
    return filtered if filtered else None


def map_format(fmt: str) -> str:
    if not fmt:
        return "text"
    if "چند گزینه" in fmt or "Multi" in fmt or "چند انتخاب" in fmt:
        return "checkbox"
    if "تک انتخاب" in fmt or fmt == "گزینه ای":
        return "radio"
    if "گزینه ای با" in fmt:
        return "checkbox"
    if "جواب بلند" in fmt or "توصیفی" in fmt:
        return "textarea"
    if "گزینه" in fmt:
        return "radio"
    return "text"


def branch_from_row(row, parent_q: str):
    opt = clean(row[4]) if len(row) > 4 else ""
    next_q = clean(row[6]) if len(row) > 6 else ""
    if not next_q and len(row) > 5:
        hint = clean(row[5])
        if hint and "؟" in hint:
            next_q = hint.split(":", 1)[-1].strip() if ":" in hint else hint
    opts = [clean(c) for c in row[7:14] if c and clean(c)]
    if not opt or not next_q:
        return None
    branch_type = "checkbox" if opts and len(opts) > 3 else ("textarea" if not opts else "radio")
    return {
        "option": opt,
        "question": next_q,
        "options": opts or None,
        "type": branch_type,
        "showIf": {"parentQuestion": parent_q, "equals": opt},
    }


def is_optional_question(question: str) -> bool:
    optional_markers = (
        "در صورت وجود",
        "اختیاری",
        "در صورت نداشتن",
    )
    return any(marker in question for marker in optional_markers)


def parse_sheet1(ws):
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        question = clean(row[0])
        if not question or question in SKIP:
            continue
        fmt = clean(row[2])
        options = split_options(clean(row[3]))
        field_type = map_format(fmt)
        if options and field_type == "text":
            field_type = "checkbox" if len(options) > 6 else "radio"
        items.append(
            {
                "question": question,
                "type": field_type,
                "options": options,
                "isAllowedEmpty": is_optional_question(question),
            }
        )
    return items


def parse_sheet_standard(ws, page_num: int):
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    items = []
    index = 0
    while index < len(rows):
        row = rows[index]
        number = clean(row[0])
        question = clean(row[1]) if len(row) > 1 else ""
        if not question or question in SKIP:
            index += 1
            continue
        if not number and not question:
            index += 1
            continue
        fmt = clean(row[3]) if len(row) > 3 else ""
        options_raw = clean(row[4]) if len(row) > 4 else ""
        field_type = map_format(fmt)
        options = split_options(options_raw)
        branches = []
        first_branch = branch_from_row(row, question)
        if first_branch:
            options = options or []
            if first_branch["option"] not in options:
                options.append(first_branch["option"])
            branches.append(first_branch)
        next_index = index + 1
        while next_index < len(rows):
            next_row = rows[next_index]
            if clean(next_row[0]) or clean(next_row[1]):
                break
            branch = branch_from_row(next_row, question)
            if branch:
                options = options or []
                if branch["option"] not in options:
                    options.append(branch["option"])
                branches.append(branch)
            next_index += 1
        if options and field_type == "text":
            field_type = "checkbox" if len(options) > 6 else "radio"
        items.append(
            {
                "question": question,
                "type": field_type,
                "options": options,
                "isAllowedEmpty": is_optional_question(question),
                "page": page_num,
            }
        )
        for branch in branches:
            items.append(
                {
                    "question": branch["question"],
                    "type": branch["type"],
                    "options": branch["options"],
                    "isAllowedEmpty": is_optional_question(branch["question"]),
                    "page": page_num,
                    "showIf": branch["showIf"],
                }
            )
        index = next_index
    return items


B2B_MODELS = [
    "B2C (Business to Consumer)",
    "B2B (Business to Business)",
    "B2B2C (Business to Business to Consumer)",
    "C2C (Consumer to Consumer)",
    "C2B (Consumer to Business)",
    "B2G (Business to Government)",
    "C2G (Consumer to Government)",
    "G2C (Government to Consumer)",
    "G2B (Government to Business)",
    "B2E (Business to Employee)",
    "P2P (Peer to Peer)",
    "G2G (Government to Government)",
]

MARKETING_TOOLS = [
    "Google Analytics",
    "Google Search Console",
    "Google Tag Manager",
    "Microsoft Clarity",
    "Hotjar",
    "Tableau",
    "Power BI",
    "Metabase",
    "CRM",
    "HubSpot",
    "Salesforce",
    "Zoho CRM",
    "Pipedrive",
    "پنل فروشگاه اینترنتی / سایت",
    "پنل شبکه‌های اجتماعی",
    "فایل اکسل / Google Sheets",
    "webengage",
    "metrics",
    "n8n",
    "adtrace",
    "matomo",
    "semrush",
    "graphana",
    "looker",
    "all seo tools",
    "VWO",
    "Optimizely",
    "AppsFlyer",
    "Adjust",
    "Hootsuite",
    "Buffer",
    "Later",
    "Sprout Social",
    "Zapier",
    "Make.com",
    "Klaviyo",
    "ActiveCampaign",
    "Mailchimp",
    "Ahrefs",
    "Moz Pro",
    "Screaming Frog SEO Spider",
    "Surfer SEO",
    "Similarweb",
    "Google Keyword Planner",
    "Mixpanel",
    "PostHog",
]

ADS_TOOLS = [
    "Google Ads",
    "Meta Ads Manager",
    "LinkedIn Campaign Manager",
    "TikTok Ads Manager",
    "X Ads",
    "یکتانت",
    "mediad",
    "صباویژن",
    "پنل تبلیغات اینفلوئنسری",
    "پنل ایمیل مارکتینگ",
    "پنل پیامک",
    "پنل نوتیفیکیشن / Push",
]

SALES_METHODS = ["آنلاین", "آفلاین", "سایر"]

PAYMENT_METHODS = [
    "درگاه پرداخت داخلی",
    "درگاه پرداخت خارجی",
    "روش های پرداخت اقساطی و تسهیلات",
    "پرداخت درب منزل",
    "کیف پول",
]

REVENUE_MODELS = [
    "فروش محصول (Product Sales)",
    "دریافت هزینه خدمات (Service Fees)",
    "اشتراک دوره‌ای، ماهانه یا سالانه (Subscription)",
    "کمیسیون از معامله (Commission)",
    "کارمزد تراکنش (Transaction Fee)",
    "درآمد از تبلیغات (Advertising)",
    "نسخه رایگان + ارتقاء پولی (Freemium)",
    "لایسنس / حق استفاده (Licensing / Royalties)",
    "اجاره یا لیزینگ (Rental / Leasing)",
    "درآمد از همکاری در فروش (Affiliate / Referral Fees)",
    "درآمدزایی از داده، گزارش، بینش آماری یا API داده (Data Monetization)",
    "اسپانسرشیپ (Sponsorship)",
    "پرداخت بر اساس مصرف (Usage / Pay-as-you-go)",
    "کارمزد واسطه‌گری (Brokerage Fees)",
    "حق عضویت (Membership Fees)",
    "هزینه راه‌اندازی، نصب یا پیاده‌سازی (Setup / Implementation Fee)",
    "هزینه نگهداری و پشتیبانی (Maintenance / Support Fees)",
    "فروش بلیت / هزینه ورودی (Ticketing / Admission Fees)",
    "کمک مالی / گرنت / حمایت بلاعوض (Donations / Grants)",
    "درآمد بهره، سود تسهیلات یا تامین مالی (Interest / Financing Income)",
    "حق فرانچایز / اعطای نمایندگی (Franchise Fees)",
]


def post_process_steps(steps: list[dict]) -> list[dict]:
    """Fix known Excel layout quirks for page 5 conditional questions."""
    cleaned = [s for s in steps if s["id"] < 39]
    business_domains = next(
        (s["options"] for s in steps if s["question"] == "حوزه اصلی کسب‌وکار شما چیست؟"),
        [],
    )

    cleaned.extend(
        [
            {
                "id": 39,
                "page": 5,
                "question": "در حال حاضر از کدام ابزارها و پنل‌ها در کسب‌وکار خود استفاده می‌کنید؟",
                "type": "checkbox",
                "options": ["ابزارهای مارکتینگی", "ابزارهای تبلیغاتی"],
            },
            {
                "id": 40,
                "page": 5,
                "question": "ابزارهای مارکتینگی خود را از لیست انتخاب کنید.",
                "type": "checkbox",
                "options": MARKETING_TOOLS,
                "showIf": {
                    "parentQuestion": "در حال حاضر از کدام ابزارها و پنل‌ها در کسب‌وکار خود استفاده می‌کنید؟",
                    "includes": "ابزارهای مارکتینگی",
                },
            },
            {
                "id": 41,
                "page": 5,
                "question": "ابزارهای تبلیغاتی خود را از لیست انتخاب کنید.",
                "type": "checkbox",
                "options": ADS_TOOLS,
                "showIf": {
                    "parentQuestion": "در حال حاضر از کدام ابزارها و پنل‌ها در کسب‌وکار خود استفاده می‌کنید؟",
                    "includes": "ابزارهای تبلیغاتی",
                },
            },
            {
                "id": 42,
                "page": 5,
                "question": "آیا قصد دارید مدل تجاری جدیدی با مشتریان خود داشته باشید؟",
                "type": "radio",
                "options": ["بله", "خیر"],
            },
            {
                "id": 43,
                "page": 5,
                "question": "مدل تجاری جدید مدنظر خود را مشخص کنید.",
                "type": "checkbox",
                "options": B2B_MODELS,
                "showIf": {
                    "parentQuestion": "آیا قصد دارید مدل تجاری جدیدی با مشتریان خود داشته باشید؟",
                    "equals": "بله",
                },
            },
            {
                "id": 44,
                "page": 5,
                "question": "آیا قصد ورود به حوزه یا حوزه‌های جدید کسب‌وکار را دارید؟",
                "type": "radio",
                "options": ["بله", "خیر"],
            },
            {
                "id": 45,
                "page": 5,
                "question": "حوزه یا حوزه‌های مدنظر خود را مشخص کنید.",
                "type": "checkbox",
                "options": business_domains,
                "showIf": {
                    "parentQuestion": "آیا قصد ورود به حوزه یا حوزه‌های جدید کسب‌وکار را دارید؟",
                    "equals": "بله",
                },
            },
            {
                "id": 46,
                "page": 5,
                "question": "آیا قصد دارید از روش جدیدی برای فروش محصول یا خدمات خود استفاده کنید؟",
                "type": "radio",
                "options": ["بله", "خیر"],
            },
            {
                "id": 47,
                "page": 5,
                "question": "روش‌های فروش مدنظر خود را مشخص کنید.",
                "type": "checkbox",
                "options": SALES_METHODS,
                "showIf": {
                    "parentQuestion": "آیا قصد دارید از روش جدیدی برای فروش محصول یا خدمات خود استفاده کنید؟",
                    "equals": "بله",
                },
            },
            {
                "id": 48,
                "page": 5,
                "question": "آیا قصد دارید از روش جدیدی برای پرداخت استفاده کنید؟",
                "type": "radio",
                "options": ["بله", "خیر"],
            },
            {
                "id": 49,
                "page": 5,
                "question": "روش‌های پرداخت مدنظر خود را مشخص کنید.",
                "type": "checkbox",
                "options": PAYMENT_METHODS,
                "showIf": {
                    "parentQuestion": "آیا قصد دارید از روش جدیدی برای پرداخت استفاده کنید؟",
                    "equals": "بله",
                },
            },
            {
                "id": 50,
                "page": 5,
                "question": "آیا قصد اضافه کردن مدل درآمدی جدید دارید؟",
                "type": "radio",
                "options": ["بله", "خیر"],
            },
            {
                "id": 51,
                "page": 5,
                "question": "مدل یا مدل‌های درآمدی مدنظر خود را انتخاب کنید.",
                "type": "checkbox",
                "options": REVENUE_MODELS,
                "showIf": {
                    "parentQuestion": "آیا قصد اضافه کردن مدل درآمدی جدید دارید؟",
                    "equals": "بله",
                },
            },
        ]
    )
    return cleaned


PAGE_LABELS = [
    "اطلاعات پایه",
    "فروش و محصولات",
    "هویت برند",
    "مخاطب و رقابت",
    "چالش‌ها و اهداف",
]

all_steps = []
step_id = 1
for page_index, sheet_name in enumerate(BRAND_SHEETS, 1):
    worksheet = wb[sheet_name]
    questions = (
        parse_sheet1(worksheet)
        if page_index == 1
        else parse_sheet_standard(worksheet, page_index)
    )
    for question in questions:
        question["id"] = step_id
        question["page"] = page_index
        all_steps.append(question)
        step_id += 1

all_steps = post_process_steps(all_steps)

lines = [
    'import type { SurveyConfig } from "./surveys";',
    "",
    "export const brandFormSurvey: SurveyConfig = {",
    '  id: "branding",',
    '  label: "افزودن اطلاعات برند",',
    '  title: "افزودن اطلاعات برند",',
    '  description: "",',
    f"  pageCount: {len(PAGE_LABELS)},",
    f"  pageLabels: {json.dumps(PAGE_LABELS, ensure_ascii=False)},",
    "  steps: [",
]

for step in all_steps:
    chunks = [
        f'    {{ id: {step["id"]}',
        f'page: {step["page"]}',
        f'question: {json.dumps(step["question"], ensure_ascii=False)}',
        f'type: "{step["type"]}"',
    ]
    if step.get("isAllowedEmpty"):
        chunks.append("isAllowedEmpty: true")
    if step.get("options"):
        chunks.append(f"options: {json.dumps(step['options'], ensure_ascii=False)}")
    if step.get("showIf"):
        chunks.append(f"showIf: {json.dumps(step['showIf'], ensure_ascii=False)}")
    lines.append(", ".join(chunks) + " },")

lines.extend(["  ],", "};", ""])
out_path.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(all_steps)} steps to {out_path}")
